from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.services.avesis.normalizer import NormalizedRecord
from app.services.faculty_catalog import Faculty
from app.services.record_filter import YearScope


RecordValue = str | int | None
RecordValueGetter = Callable[[NormalizedRecord], RecordValue]


@dataclass(frozen=True)
class ColumnDefinition:
    header: str
    get_value: RecordValueGetter
    is_hyperlink: bool = False


@dataclass(frozen=True)
class SheetDefinition:
    sheet_name: str
    table_name: str
    columns: tuple[ColumnDefinition, ...]


RECORD_TYPE_ORDER = (
    "article",
    "conference_paper",
    "book",
    "project",
    "patent",
)

RECORD_TYPE_LABELS = {
    "article": "Makale",
    "conference_paper": "Bildiri",
    "book": "Kitap",
    "project": "Proje",
    "patent": "Patent",
}

SHEET_DEFINITIONS = {
    "article": SheetDefinition(
        sheet_name="01_Makaleler",
        table_name="ArticlesTable",
        columns=(
            ColumnDefinition(
                "Akademisyen",
                lambda record: record.academician_name,
            ),
            ColumnDefinition("Yıl", lambda record: record.year),
            ColumnDefinition("Makale Adı", lambda record: record.title),
            ColumnDefinition(
                "Yazarlar",
                lambda record: record.contributor_names,
            ),
            ColumnDefinition(
                "Dergi",
                lambda record: record.data.get("journal_name"),
            ),
            ColumnDefinition(
                "Cilt",
                lambda record: record.data.get("volume"),
            ),
            ColumnDefinition(
                "Sayı",
                lambda record: record.data.get("issue"),
            ),
            ColumnDefinition(
                "Sayfalar",
                lambda record: record.data.get("pages"),
            ),
            ColumnDefinition(
                "DOI",
                lambda record: record.data.get("doi"),
            ),
            ColumnDefinition(
                "İndeksler",
                lambda record: record.data.get("journal_indexes"),
            ),
            ColumnDefinition(
                "AVESİS'te Aç",
                lambda record: record.source_url,
                is_hyperlink=True,
            ),
        ),
    ),
    "conference_paper": SheetDefinition(
        sheet_name="02_Bildiriler",
        table_name="ConferencePapersTable",
        columns=(
            ColumnDefinition(
                "Akademisyen",
                lambda record: record.academician_name,
            ),
            ColumnDefinition("Yıl", lambda record: record.year),
            ColumnDefinition("Bildiri Adı", lambda record: record.title),
            ColumnDefinition(
                "Yazarlar",
                lambda record: record.contributor_names,
            ),
            ColumnDefinition(
                "Konferans",
                lambda record: record.data.get("conference_name"),
            ),
            ColumnDefinition(
                "Tarih",
                lambda record: record.data.get("conference_date"),
            ),
            ColumnDefinition(
                "Şehir",
                lambda record: record.data.get("city"),
            ),
            ColumnDefinition(
                "Ülke",
                lambda record: record.data.get("country"),
            ),
            ColumnDefinition(
                "Sayfalar",
                lambda record: record.data.get("pages"),
            ),
            ColumnDefinition(
                "DOI",
                lambda record: record.data.get("doi"),
            ),
            ColumnDefinition(
                "AVESİS'te Aç",
                lambda record: record.source_url,
                is_hyperlink=True,
            ),
        ),
    ),
    "book": SheetDefinition(
        sheet_name="03_Kitaplar",
        table_name="BooksTable",
        columns=(
            ColumnDefinition(
                "Akademisyen",
                lambda record: record.academician_name,
            ),
            ColumnDefinition("Yıl", lambda record: record.year),
            ColumnDefinition("Başlık", lambda record: record.title),
            ColumnDefinition(
                "Yazarlar",
                lambda record: record.contributor_names,
            ),
            ColumnDefinition(
                "Yayın Türü",
                lambda record: record.data.get("publication_type"),
            ),
            ColumnDefinition(
                "Ana Kitap Adı",
                lambda record: record.data.get("book_title"),
            ),
            ColumnDefinition(
                "Yayınevi",
                lambda record: record.data.get("publisher"),
            ),
            ColumnDefinition(
                "Şehir",
                lambda record: record.data.get("city"),
            ),
            ColumnDefinition(
                "Sayfalar",
                lambda record: record.data.get("pages"),
            ),
            ColumnDefinition(
                "Editörler",
                lambda record: record.data.get("editors"),
            ),
            ColumnDefinition(
                "AVESİS'te Aç",
                lambda record: record.source_url,
                is_hyperlink=True,
            ),
        ),
    ),
    "project": SheetDefinition(
        sheet_name="04_Projeler",
        table_name="ProjectsTable",
        columns=(
            ColumnDefinition(
                "Akademisyen",
                lambda record: record.academician_name,
            ),
            ColumnDefinition("Proje Adı", lambda record: record.title),
            ColumnDefinition(
                "Proje Ekibi/Roller",
                lambda record: record.contributor_text,
            ),
            ColumnDefinition(
                "Proje Türü",
                lambda record: record.data.get("project_type"),
            ),
            ColumnDefinition(
                "Destek Programı",
                lambda record: record.data.get("support_program"),
            ),
            ColumnDefinition(
                "Destekleyen Kuruluş",
                lambda record: record.data.get(
                    "supporting_organization"
                ),
            ),
            ColumnDefinition(
                "Başlangıç",
                lambda record: record.data.get("start_date"),
            ),
            ColumnDefinition(
                "Bitiş",
                lambda record: record.data.get("end_date"),
            ),
            ColumnDefinition(
                "AVESİS'te Aç",
                lambda record: record.source_url,
                is_hyperlink=True,
            ),
        ),
    ),
    "patent": SheetDefinition(
        sheet_name="05_Patentler",
        table_name="PatentsTable",
        columns=(
            ColumnDefinition(
                "Akademisyen",
                lambda record: record.academician_name,
            ),
            ColumnDefinition("Patent Adı", lambda record: record.title),
            ColumnDefinition(
                "Mucitler",
                lambda record: record.contributor_names,
            ),
            ColumnDefinition(
                "Fikri Mülkiyet",
                lambda record: record.data.get(
                    "intellectual_property"
                ),
            ),
            ColumnDefinition(
                "Patent Sınıfı",
                lambda record: record.data.get("patent_class"),
            ),
            ColumnDefinition(
                "Tescil No",
                lambda record: record.data.get(
                    "registration_number"
                ),
            ),
            ColumnDefinition(
                "Tescil Tipi",
                lambda record: record.data.get(
                    "registration_type"
                ),
            ),
            ColumnDefinition(
                "Başvuru Ülkesi/Kuruluşu",
                lambda record: record.data.get(
                    "application_country"
                ),
            ),
            ColumnDefinition(
                "Başvuru Tarihi",
                lambda record: record.data.get("application_date"),
            ),
            ColumnDefinition(
                "Tescil Tarihi",
                lambda record: record.data.get(
                    "registration_date"
                ),
            ),
            ColumnDefinition(
                "Durum",
                lambda record: record.data.get("status"),
            ),
            ColumnDefinition(
                "AVESİS'te Aç",
                lambda record: record.source_url,
                is_hyperlink=True,
            ),
        ),
    ),
}


class ExcelExporter:
    _header_fill = PatternFill(
        fill_type="solid",
        fgColor="003B71",
    )
    _header_font = Font(color="FFFFFF", bold=True)
    _hyperlink_font = Font(color="0563C1", underline="single")

    def export(
        self,
        records: Iterable[NormalizedRecord],
        academicians: Iterable[Faculty],
        selected_record_types: set[str],
        year_scope: YearScope,
        output_path: Path,
    ) -> Path:
        record_list = list(records)
        academician_list = list(academicians)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "00_Ozet"

        self._write_summary_sheet(
            sheet=summary_sheet,
            records=record_list,
            academicians=academician_list,
            selected_record_types=selected_record_types,
            year_scope=year_scope,
        )

        for record_type in RECORD_TYPE_ORDER:
            if record_type not in selected_record_types:
                continue

            definition = SHEET_DEFINITIONS[record_type]
            sheet = workbook.create_sheet(definition.sheet_name)

            type_records = [
                record
                for record in record_list
                if record.record_type == record_type
            ]

            self._write_record_sheet(
                sheet=sheet,
                definition=definition,
                records=type_records,
            )

        workbook.save(output_path)
        workbook.close()

        return output_path.resolve()

    def _write_summary_sheet(
        self,
        sheet,
        records: list[NormalizedRecord],
        academicians: list[Faculty],
        selected_record_types: set[str],
        year_scope: YearScope,
    ) -> None:
        selected_types = [
            record_type
            for record_type in RECORD_TYPE_ORDER
            if record_type in selected_record_types
        ]

        headers = [
            "Akademisyen",
            "Birim",
            *[
                RECORD_TYPE_LABELS[record_type]
                for record_type in selected_types
            ],
            "Toplam",
            "Tarihi Belirsiz",
        ]

        last_column = get_column_letter(len(headers))

        sheet.merge_cells(f"A1:{last_column}1")
        sheet["A1"] = "DEÜ AVESİS Akademik Rapor"
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = self._header_fill
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet["A3"] = "Zaman Kapsamı"
        sheet["B3"] = self._format_year_scope(year_scope)

        sheet["A4"] = "Rapor Oluşturma Zamanı"
        sheet["B4"] = datetime.now().strftime("%d.%m.%Y %H:%M")

        sheet["A5"] = "Seçilen Kayıt Türleri"
        sheet["B5"] = ", ".join(
            RECORD_TYPE_LABELS[record_type]
            for record_type in selected_types
        )

        header_row = 7

        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )
            self._style_header_cell(cell)

        for row_index, academician in enumerate(
            academicians,
            start=header_row + 1,
        ):
            academician_records = [
                record
                for record in records
                if record.academician_id == academician.id
            ]

            values: list[RecordValue] = [
                academician.full_name,
                academician.unit,
            ]

            for record_type in selected_types:
                count = sum(
                    record.record_type == record_type
                    for record in academician_records
                )
                values.append(count)

            values.append(len(academician_records))
            values.append(
                sum(
                    record.year is None
                    for record in academician_records
                )
            )

            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )
                self._style_data_cell(cell)

        if academicians:
            table = Table(
                displayName="SummaryTable",
                ref=f"A{header_row}:{last_column}{sheet.max_row}",
            )
            table.tableStyleInfo = self._table_style()
            sheet.add_table(table)

        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.sheet_view.showGridLines = False
        self._adjust_column_widths(sheet)

    def _write_record_sheet(
        self,
        sheet,
        definition: SheetDefinition,
        records: list[NormalizedRecord],
    ) -> None:
        for column_index, column in enumerate(
            definition.columns,
            start=1,
        ):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=column.header,
            )
            self._style_header_cell(cell)

        sorted_records = sorted(
            records,
            key=lambda record: (
                -(record.year or 0),
                record.title.casefold(),
            ),
        )

        for row_index, record in enumerate(sorted_records, start=2):
            for column_index, column in enumerate(
                definition.columns,
                start=1,
            ):
                value = column.get_value(record)
                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                )

                if column.is_hyperlink and value:
                    cell.value = "Aç"
                    cell.hyperlink = str(value)
                    cell.font = self._hyperlink_font
                else:
                    cell.value = value

                self._style_data_cell(cell)

        if records:
            last_column = get_column_letter(len(definition.columns))
            table = Table(
                displayName=definition.table_name,
                ref=f"A1:{last_column}{sheet.max_row}",
            )
            table.tableStyleInfo = self._table_style()
            sheet.add_table(table)

        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        self._adjust_column_widths(sheet)

    def _style_header_cell(self, cell) -> None:
        cell.fill = self._header_fill
        cell.font = self._header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    @staticmethod
    def _style_data_cell(cell) -> None:
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    @staticmethod
    def _table_style() -> TableStyleInfo:
        return TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

    @staticmethod
    def _format_year_scope(year_scope: YearScope) -> str:
        if year_scope.is_all_years:
            return "Tüm yıllar"

        if year_scope.start_year == year_scope.end_year:
            return str(year_scope.start_year)

        return f"{year_scope.start_year} - {year_scope.end_year}"

    @staticmethod
    def _adjust_column_widths(sheet) -> None:
        for column_cells in sheet.iter_cols():
            maximum_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            column_letter = get_column_letter(
                column_cells[0].column
            )

            sheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                55,
            )