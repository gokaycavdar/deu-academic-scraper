from openpyxl import load_workbook

from app.services.avesis.normalizer import NormalizedRecord
from app.services.excel_exporter import ExcelExporter
from app.services.faculty_catalog import Faculty
from app.services.record_filter import YearScope


def test_exports_summary_and_selected_record_sheets(
    tmp_path,
) -> None:
    academician = Faculty(
        id="gokhan-dalkilic",
        sort_order=10,
        full_name="Gökhan Dalkılıç",
        unit="Mühendislik Fakültesi / Bilgisayar Mühendisliği",
        profile_url="https://avesis.deu.edu.tr/gokhan.dalkilic",
    )

    article = NormalizedRecord(
        academician_id=academician.id,
        academician_name=academician.full_name,
        record_id="article-test",
        record_type="article",
        title="Test Makalesi",
        contributor_names="Yazar A., Yazar B.",
        contributor_text="Yazar A., Yazar B.",
        source_url="https://avesis.deu.edu.tr/yayin/article-test",
        citation_text="Test Dergisi, 2026",
        year=2026,
        data={
            "journal_name": "Test Dergisi",
            "volume": "15",
            "issue": "2",
            "pages": "ss.10-20",
            "doi": "10.1000/test",
            "journal_indexes": "Scopus",
            "scope": "Uluslararası",
            "peer_review": "Hakemli",
            "issn": "1234-5678",
        },
        source_names=("AVESİS", "YÖK Akademik"),
    )

    patent = NormalizedRecord(
        academician_id=academician.id,
        academician_name=academician.full_name,
        record_id="patent-test",
        record_type="patent",
        title="Test Patenti",
        contributor_names="Mucit A.",
        contributor_text="Mucit A.",
        source_url=(
            "https://avesis.deu.edu.tr/"
            "fikrimulkiyet/patent-test"
        ),
        citation_text="Patent, BÖLÜM G Fizik",
        year=2024,
        data={
            "intellectual_property": "Patent",
            "patent_class": "BÖLÜM G Fizik",
            "registration_number": "2024 00001",
            "registration_type": "Standart Tescil",
            "application_country": "Türkiye",
            "application_date": "01.01.2023",
            "registration_date": "01.01.2024",
            "status": "Tescil Edildi",
            "applicants": "Dokuz Eylül Üniversitesi",
        },
    )

    output_path = tmp_path / "akademik_rapor.xlsx"

    ExcelExporter().export(
        records=[article, patent],
        academicians=[academician],
        selected_record_types={"article", "patent"},
        year_scope=YearScope.all_years(),
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    try:
        assert workbook.sheetnames == [
            "00_Ozet",
            "01_Makaleler",
            "05_Patentler",
        ]

        summary_sheet = workbook["00_Ozet"]

        assert summary_sheet["A1"].value == "DEÜ Akademik Rapor"
        assert summary_sheet["C8"].value == 1
        assert summary_sheet["D8"].value == 1
        assert summary_sheet["E8"].value == 2
        assert summary_sheet["F8"].value == 0

        article_sheet = workbook["01_Makaleler"]
        article_columns = _header_columns(article_sheet)

        assert _cell_value(
            article_sheet,
            article_columns,
            "Makale Adı",
        ) == "Test Makalesi"
        assert _cell_value(
            article_sheet,
            article_columns,
            "Kayıt Kaynağı",
        ) == "AVESİS + YÖK Akademik"
        assert _cell_value(
            article_sheet,
            article_columns,
            "Dergi",
        ) == "Test Dergisi"
        assert _cell_value(
            article_sheet,
            article_columns,
            "DOI",
        ) == "10.1000/test"
        assert _cell_value(
            article_sheet,
            article_columns,
            "Kapsam",
        ) == "Uluslararası"
        assert _cell_value(
            article_sheet,
            article_columns,
            "Hakem Durumu",
        ) == "Hakemli"
        assert _cell_value(
            article_sheet,
            article_columns,
            "ISSN",
        ) == "1234-5678"

        article_link_cell = article_sheet.cell(
            row=2,
            column=article_columns["AVESİS'te Aç"],
        )

        assert article_link_cell.value == "Aç"
        assert article_link_cell.hyperlink is not None
        assert article_link_cell.hyperlink.target == (
            "https://avesis.deu.edu.tr/yayin/article-test"
        )

        patent_sheet = workbook["05_Patentler"]
        patent_columns = _header_columns(patent_sheet)

        assert _cell_value(
            patent_sheet,
            patent_columns,
            "Patent Adı",
        ) == "Test Patenti"
        assert _cell_value(
            patent_sheet,
            patent_columns,
            "Patent Başvuru Sahibi",
        ) == "Dokuz Eylül Üniversitesi"
        assert _cell_value(
            patent_sheet,
            patent_columns,
            "Durum",
        ) == "Tescil Edildi"
    finally:
        workbook.close()


def _header_columns(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def _cell_value(
    sheet,
    columns: dict[str, int],
    header: str,
):
    return sheet.cell(row=2, column=columns[header]).value